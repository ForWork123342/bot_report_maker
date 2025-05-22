import io
from aiogram import Router, F
from aiogram.types import CallbackQuery, BufferedInputFile
from aiogram.fsm.context import FSMContext
from models.report_states import NightReportForm
import openpyxl
from openpyxl.styles import Font, PatternFill

from datetime import date


from services.send_night_report import send_night_report as night_report

router = Router(name=__name__)

@router.callback_query(F.data.startswith('finish_night_report'), NightReportForm.end_report)
async def send_night_report(callback: CallbackQuery, state: FSMContext):
    data = await state.get_data()
    report_data = night_report(data["report_data"])



    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "report_data"

    headers = ["Дата", "ФИО", "Товар", "Тара", "Вес", "Вес тары", "Вес содержимого"]
    ws.append(headers)

    # Задаем стиль для заголовков
    header_font = Font(bold=True, color="FFFFFF")  # Белый жирный шрифт
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")  # Синий фон

    # Применяем стиль к заголовкам
    for cell in ws[1]:  # Первая строка (заголовки)
        cell.font = header_font
        cell.fill = header_fill

    for row in report_data:
        ws.append(row)

    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)

    # Создаем BufferedInputFile из буфера
    file = BufferedInputFile(
        file=excel_buffer.getvalue(),
        filename=f"{date.today().isoformat()}_night_report.xlsx"
    )

    await callback.message.delete()
    await state.clear()

    data_after_finish = await state.get_data()
    print(data_after_finish)  

    await callback.message.answer_document(
        document=file,
        caption="Отчет успешно отправлен ☑️\nПоследний отчет:"
    )